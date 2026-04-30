"""视频记录管理 API。"""

import asyncio
import time
import random
import json
from typing import List
from datetime import datetime
from fastapi import APIRouter, HTTPException, File, UploadFile
from sse_starlette.sse import EventSourceResponse

from app.repositories.db import get_db
from app.schemas.video_record import VideoRecordCreate, VideoRecordUpdate, VideoRecordResponse
from app.integrations.yt_dlp_client import downloader
from app.integrations.douyin_client import douyin_parser, is_douyin_url


router = APIRouter()

# 全局字典：存储每个记录的 AI 分析更新队列
ai_update_queues = {}

# 全局字典：存储后台任务的进度事件队列
background_task_queues = {}

# 从配置文件读取并发数
def get_max_concurrency():
    """从配置文件读取最大并发数。"""
    try:
        import yaml
        from pathlib import Path
        config_path = Path(__file__).parent.parent.parent.parent / "api_config.yaml"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
                return config.get("max_concurrency", 5)
    except Exception as e:
        print(f"[并发控制] 读取配置失败: {e}，使用默认值5")
    return 5

# 全局信号量：限制并发处理的视频数量
video_processing_semaphore = asyncio.Semaphore(get_max_concurrency())
print(f"[并发控制] 初始化完成，最大并发数: {get_max_concurrency()}")


# 独立的后台视频处理函数（不依赖 SSE 连接）
async def process_video_background(video_url: str, sheet_id: str, idx: int, total: int, task_id: str):
    """后台处理视频，即使 SSE 断开也会继续执行。使用信号量控制并发。"""

    def send_event(event_type: str, data: dict):
        """发送事件到队列（如果队列存在）。"""
        if task_id in background_task_queues:
            try:
                background_task_queues[task_id].put_nowait({
                    "event": event_type,
                    "data": json.dumps(data, ensure_ascii=False)
                })
            except:
                pass  # 队列已满或已关闭，忽略

    # 获取信号量，限制并发数量
    async with video_processing_semaphore:
        max_concurrency = get_max_concurrency()
        current_concurrency = max_concurrency - video_processing_semaphore._value + 1  # +1因为当前任务已获取信号量
        print(f"[视频记录] 开始处理视频 {idx}/{total}（当前并发：{current_concurrency}/{max_concurrency}）")
        loop = asyncio.get_event_loop()
        record_id = None

        try:
            # 创建空记录
            with get_db() as conn:
                cursor = conn.execute(
                    """INSERT INTO video_records (video_url, sheet_id) VALUES (?, ?)""",
                    (video_url, sheet_id)
                )
                record_id = cursor.lastrowid

                row = conn.execute(
                    "SELECT * FROM video_records WHERE id = ?", (record_id,)
                ).fetchone()

                send_event("created", dict(row))

            # 解析视频信息
            send_event("status", {"message": f"正在解析视频 {idx}/{total}..."})

            if is_douyin_url(video_url):
                video_info = await loop.run_in_executor(None, douyin_parser.parse, video_url)
            else:
                video_info = await loop.run_in_executor(None, downloader.parse_video, video_url)

            # 提取基本信息
            video_time = None
            if video_info.get("upload_date"):
                try:
                    upload_date = video_info["upload_date"]
                    video_time = datetime.strptime(upload_date, "%Y%m%d").isoformat()
                except Exception:
                    pass

            video_copy = video_info.get("description", "")[:500] if video_info.get("description") else ""
            if not video_copy:
                video_copy = video_info.get("title", "")[:500]

            likes = video_info.get("like_count", 0) or 0
            comments = video_info.get("comment_count", 0) or 0
            shares = video_info.get("share_count", 0) or 0
            collects = video_info.get("collect_count", 0) or 0

            # 更新数据库并推送基本信息
            with get_db() as conn:
                conn.execute(
                    """UPDATE video_records
                       SET video_time = ?, video_copy = ?, likes = ?, comments = ?, shares = ?, collects = ?, updated_at = datetime('now')
                       WHERE id = ?""",
                    (video_time, video_copy, likes, comments, shares, collects, record_id)
                )

                row = conn.execute(
                    "SELECT * FROM video_records WHERE id = ?", (record_id,)
                ).fetchone()

                send_event("basic_info", dict(row))

            # 下载视频
            send_event("status", {"message": f"正在下载视频 {idx}/{total}..."})

            video_file_path = ""
            try:
                formats = video_info.get("formats", [])
                if formats and len(formats) > 0:
                    # 尝试所有可用的格式（优先无水印，失败后尝试有水印）
                    download_success = False
                    last_error = None

                    for format_idx, fmt in enumerate(formats):
                        direct_url = fmt.get("_direct_url") or fmt.get("url")
                        if not direct_url:
                            continue

                        format_label = fmt.get("label", f"格式 {format_idx + 1}")
                        print(f"[视频下载] 尝试下载: {format_label}")

                        try:
                            import requests
                            from pathlib import Path

                            title = video_info.get("title", "video")
                            safe_title = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in title)[:60]
                            filename = f"{safe_title}_{record_id}.mp4"

                            download_dir = Path(__file__).parent.parent.parent.parent / "downloads"
                            download_dir.mkdir(parents=True, exist_ok=True)
                            filepath = download_dir / filename

                            await loop.run_in_executor(None, download_with_retry, direct_url, filepath, 2)
                            video_file_path = str(filepath)
                            download_success = True
                            print(f"[视频下载] ✓ {format_label} 下载成功")
                            break  # 下载成功，跳出循环

                        except Exception as e:
                            last_error = e
                            print(f"[视频下载] ✗ {format_label} 下载失败: {e}")
                            # 继续尝试下一个格式

                    if not download_success:
                        # 所有格式都失败，尝试使用浏览器下载插件
                        print(f"[视频下载] 常规方法全部失败，启动浏览器下载插件...")
                        try:
                            from app.integrations.browser_downloader import download_with_browser

                            # 准备文件路径
                            title = video_info.get("title", "video")
                            safe_title = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in title)[:60]
                            filename = f"{safe_title}_{record_id}.mp4"

                            from pathlib import Path
                            download_dir = Path(__file__).parent.parent.parent.parent / "downloads"
                            download_dir.mkdir(parents=True, exist_ok=True)
                            filepath = download_dir / filename

                            # 使用浏览器下载
                            browser_success = await loop.run_in_executor(
                                None,
                                download_with_browser,
                                video_url,  # 使用原始视频页面URL
                                str(filepath)
                            )

                            if browser_success:
                                video_file_path = str(filepath)
                                download_success = True
                                print(f"[视频下载] ✓ 浏览器插件下载成功")
                                print(f"[视频下载] video_file_path = {video_file_path}")
                                print(f"[视频下载] download_success = {download_success}")
                            else:
                                print(f"[视频下载] ✗ 浏览器插件也失败")

                        except ImportError as ie:
                            print(f"[视频下载] ✗ 浏览器插件未安装: {ie}")
                            print(f"[视频下载] 提示: 请运行 'pip install playwright' 和 'playwright install chromium'")
                        except Exception as browser_error:
                            print(f"[视频下载] ✗ 浏览器插件异常: {browser_error}")
                            import traceback
                            traceback.print_exc()

                        # 如果浏览器方法也失败，标记为解析失败
                        if not download_success:
                            print(f"[视频下载] 所有下载方法均失败，标记为解析失败")
                            with get_db() as conn:
                                # 在所有文本字段填写失败提示
                                fail_message = "视频解析失败，请手动上传视频到视频播放栏"
                                conn.execute(
                                    """UPDATE video_records
                                       SET category = ?, product = ?, golden_3s_copy = ?,
                                           transcript = ?, viral_analysis = ?, scene_analysis = ?,
                                           updated_at = datetime('now')
                                       WHERE id = ?""",
                                    (fail_message, fail_message, fail_message,
                                     fail_message, fail_message, fail_message, record_id)
                                )

                                row = conn.execute(
                                    "SELECT * FROM video_records WHERE id = ?", (record_id,)
                                ).fetchone()

                                send_event("video_failed", dict(row))

                            # 跳过AI分析，直接返回
                            return record_id

                    # 更新视频文件路径
                    with get_db() as conn:
                        conn.execute(
                            "UPDATE video_records SET video_file_path = ?, updated_at = datetime('now') WHERE id = ?",
                            (video_file_path, record_id)
                        )

                        row = conn.execute(
                            "SELECT * FROM video_records WHERE id = ?", (record_id,)
                        ).fetchone()

                        send_event("video_downloaded", dict(row))

                        # 启动后台 AI 分析任务
                        send_event("status", {"message": f"AI 分析已启动 {idx}/{total}..."})

                        from app.services.video_compressor import compressor
                        from app.services.gemini_video_analyzer import analyzer

                        def compress_and_analyze(vid_path: str, rec_id: int):
                            try:
                                import os
                                from app.repositories.db import get_db

                                print(f"[压缩分析] ========== 开始处理记录 {rec_id} ==========", flush=True)
                                print(f"[视频记录 {rec_id}] 开始压缩视频: {os.path.basename(vid_path)}", flush=True)
                                compress_result = compressor.compress_video(vid_path, "medium", 1280)

                                if compress_result:
                                    print(f"[视频记录 {rec_id}] 压缩完成，开始 AI 分析", flush=True)
                                else:
                                    print(f"[视频记录 {rec_id}] 压缩失败，使用原视频进行 AI 分析", flush=True)

                                if analyzer.is_available():
                                    print(f"[视频记录 {rec_id}] 开始 AI 分析视频", flush=True)

                                    # 获取当前工作表的预设
                                    preset = None
                                    with get_db() as conn:
                                        preset_row = conn.execute(
                                            "SELECT product_name, product_description, product_image_paths FROM sheet_presets WHERE sheet_id = ?",
                                            (sheet_id,)
                                        ).fetchone()

                                        if preset_row:
                                            preset = {
                                                "product_name": preset_row["product_name"],
                                                "product_description": preset_row["product_description"],
                                                "product_image_paths": preset_row["product_image_paths"]
                                            }
                                            print(f"[视频记录 {rec_id}] 找到预设配置:", flush=True)
                                            print(f"  - 产品名称: {preset['product_name']}", flush=True)
                                            print(f"  - 产品说明: {preset['product_description']}", flush=True)
                                            print(f"  - 图片路径: {preset['product_image_paths']}", flush=True)

                                    analysis_result = analyzer.analyze_compressed_video(vid_path, preset)

                                    if analysis_result:
                                        # 检查是否是错误结果
                                        if analysis_result.get("_error"):
                                            print(f"[视频记录 {rec_id}] AI API 错误: {analysis_result.get('error_message')}", flush=True)

                                            # 推送错误到 SSE 队列
                                            if rec_id in ai_update_queues:
                                                try:
                                                    error_data = {
                                                        "_error": True,
                                                        "error_type": analysis_result.get("error_type", "unknown"),
                                                        "error_message": analysis_result.get("error_message", "AI API 暂时不可用")
                                                    }
                                                    ai_update_queues[rec_id].put_nowait(error_data)
                                                    print(f"[SSE推送] ✓ 已推送 AI 错误通知到队列", flush=True)
                                                except Exception as e:
                                                    print(f"[SSE推送] ✗ 推送错误通知失败: {e}", flush=True)
                                        else:
                                            print(f"[视频记录 {rec_id}] AI 分析完成，更新数据库", flush=True)
                                            with get_db() as conn:
                                                category = analysis_result.get("category", "")
                                                product = analysis_result.get("product", "")
                                                golden_3s = analysis_result.get("golden_3s", "")
                                                transcript = analysis_result.get("transcript", "")
                                                viral_analysis = analysis_result.get("viral_analysis", "")
                                                scenes = analysis_result.get("scenes", "")
                                                copywriting = analysis_result.get("copywriting", "")

                                                conn.execute(
                                                    """UPDATE video_records
                                                       SET category = ?, product = ?, golden_3s_copy = ?, transcript = ?, viral_analysis = ?, scene_analysis = ?, copywriting = ?, updated_at = datetime('now')
                                                       WHERE id = ?""",
                                                    (category, product, golden_3s, transcript, viral_analysis, scenes, copywriting, rec_id)
                                                )
                                            print(f"[视频记录 {rec_id}] AI 分析结果已保存到数据库", flush=True)

                                            # 推送到 SSE 队列
                                            print(f"[SSE推送] 检查记录 {rec_id} 的队列是否存在...", flush=True)
                                            print(f"[SSE推送] 当前队列列表: {list(ai_update_queues.keys())}", flush=True)

                                            if rec_id in ai_update_queues:
                                                try:
                                                    update_data = {
                                                        "category": category,
                                                        "product": product,
                                                        "golden_3s_copy": golden_3s,
                                                        "transcript": transcript,
                                                        "viral_analysis": viral_analysis,
                                                        "scene_analysis": scenes,
                                                        "copywriting": copywriting
                                                    }
                                                    ai_update_queues[rec_id].put_nowait(update_data)
                                                    print(f"[SSE推送] ✓ 成功推送记录 {rec_id} 的 AI 分析结果到队列", flush=True)
                                                    print(f"[SSE推送] 数据预览: golden_3s={golden_3s[:50]}...", flush=True)
                                                except Exception as e:
                                                    print(f"[SSE推送] ✗ 记录 {rec_id} 推送失败: {e}", flush=True)
                                            else:
                                                print(f"[SSE推送] ✗ 记录 {rec_id} 的队列不存在，可能SSE连接未建立", flush=True)
                                    else:
                                        print(f"[视频记录 {rec_id}] AI 分析失败")
                                else:
                                    print(f"[视频记录 {rec_id}] Gemini 服务不可用，跳过 AI 分析")

                            except Exception as e:
                                print(f"[视频记录 {rec_id}] 压缩和分析任务失败: {e}")
                                import traceback
                                traceback.print_exc()

                        loop.run_in_executor(None, compress_and_analyze, video_file_path, record_id)

            except Exception as e:
                print(f"[视频记录] 视频 {idx} 下载失败: {e}")
                send_event("error", {"message": f"视频 {idx} 下载失败: {str(e)}"})

        except Exception as e:
            print(f"[视频记录] 视频 {idx} 处理失败: {e}")
            send_event("error", {"message": f"视频 {idx} 处理失败: {str(e)}"})
        finally:
            print(f"[视频记录] 视频 {idx}/{total} 处理完成，释放并发槽位")

        return record_id


def download_with_retry(url: str, filepath, max_retries: int = 3):
    """
    带智能重试机制的视频下载函数，模拟真实用户行为以避免反爬虫识别。

    策略：
    - 随机延迟（模拟人类思考时间）
    - 指数退避（每次失败后等待时间递增）
    - User-Agent 轮换
    - SSL 错误专门处理
    """
    import requests

    # 多个真实浏览器 User-Agent 池
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15",
    ]

    for attempt in range(max_retries):
        try:
            # 第一次请求前添加随机延迟（模拟用户浏览行为）
            if attempt > 0:
                # 指数退避：1-3秒、2-5秒、4-8秒
                base_delay = 2 ** attempt
                jitter = random.uniform(0, base_delay)
                wait_time = base_delay + jitter
                print(f"[下载重试] 第 {attempt + 1} 次重试，等待 {wait_time:.1f} 秒...")
                time.sleep(wait_time)
            else:
                # 首次请求也添加小延迟（0.5-1.5秒），模拟真实用户
                time.sleep(random.uniform(0.5, 1.5))

            # 随机选择 User-Agent
            headers = {
                "User-Agent": random.choice(user_agents),
                "Referer": "https://www.douyin.com/",
                "Accept": "video/webm,video/ogg,video/*;q=0.9,application/ogg;q=0.7,audio/*;q=0.6,*/*;q=0.5",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
            }

            # 发起请求，增加超时时间
            response = requests.get(
                url,
                headers=headers,
                stream=True,
                timeout=(10, 120),  # (连接超时, 读取超时)
                verify=True  # 启用 SSL 验证
            )
            response.raise_for_status()

            # 下载文件
            downloaded_size = 0
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded_size += len(chunk)

            print(f"[下载成功] 已下载 {downloaded_size / 1024 / 1024:.2f} MB")
            return True

        except requests.exceptions.SSLError as e:
            print(f"[下载失败] SSL 错误 (尝试 {attempt + 1}/{max_retries}): {str(e)[:100]}")
            if attempt == max_retries - 1:
                raise Exception(f"SSL 连接失败，已重试 {max_retries} 次")

        except requests.exceptions.Timeout as e:
            print(f"[下载失败] 超时 (尝试 {attempt + 1}/{max_retries}): {str(e)[:100]}")
            if attempt == max_retries - 1:
                raise Exception(f"下载超时，已重试 {max_retries} 次")

        except requests.exceptions.ConnectionError as e:
            print(f"[下载失败] 连接错误 (尝试 {attempt + 1}/{max_retries}): {str(e)[:100]}")
            if attempt == max_retries - 1:
                raise Exception(f"网络连接失败，已重试 {max_retries} 次")

        except Exception as e:
            print(f"[下载失败] 未知错误 (尝试 {attempt + 1}/{max_retries}): {str(e)[:100]}")
            if attempt == max_retries - 1:
                raise

    return False


@router.get("/api/video-records", response_model=List[VideoRecordResponse])
async def get_video_records(sheet_id: str = None):
    """获取视频记录，可按工作表筛选。"""
    with get_db() as conn:
        if sheet_id:
            rows = conn.execute(
                "SELECT * FROM video_records WHERE sheet_id = ? ORDER BY id DESC",
                (sheet_id,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM video_records ORDER BY id DESC"
            ).fetchall()
        return [dict(row) for row in rows]


@router.post("/api/video-records")
async def create_video_record(req: VideoRecordCreate):
    """创建视频记录并流式推送进度更新。支持批量创建多个视频。"""
    from sse_starlette.sse import EventSourceResponse
    from app.utils.url_extractor import extract_all_urls, is_video_url
    import uuid

    # 生成唯一的任务 ID
    task_id = str(uuid.uuid4())

    async def event_generator():
        try:
            # 1. 提取所有视频链接
            all_urls = extract_all_urls(req.video_url)
            video_urls = [url for url in all_urls if is_video_url(url)]

            if not video_urls:
                video_urls = [req.video_url]

            print(f"[视频记录] 识别到 {len(video_urls)} 个视频链接，开始并行处理")

            # 2. 创建事件队列
            event_queue = asyncio.Queue()
            background_task_queues[task_id] = event_queue

            # 3. 启动所有后台任务（不依赖 SSE 连接）
            tasks = []
            for idx, url in enumerate(video_urls, 1):
                task = asyncio.create_task(
                    process_video_background(url, req.sheet_id, idx, len(video_urls), task_id)
                )
                tasks.append(task)
                print(f"[视频记录] 已启动后台任务 {idx}/{len(video_urls)}")

            # 4. 从队列中读取事件并推送（SSE 断开不影响后台任务）
            completed = 0
            while completed < len(tasks):
                try:
                    # 等待事件
                    event = await asyncio.wait_for(event_queue.get(), timeout=0.1)
                    yield event
                except asyncio.TimeoutError:
                    # 检查是否有任务完成
                    for task in tasks:
                        if task.done() and not hasattr(task, '_counted'):
                            completed += 1
                            task._counted = True
                            if task.exception():
                                print(f"[视频记录] 任务异常: {task.exception()}")

            # 5. 清空队列中剩余的事件
            while not event_queue.empty():
                yield await event_queue.get()

        except Exception as e:
            print(f"[视频记录] 批量创建失败: {e}")
            yield {
                "event": "error",
                "data": json.dumps({"message": f"批量创建失败: {str(e)}"}, ensure_ascii=False)
            }
        finally:
            # 清理队列
            if task_id in background_task_queues:
                del background_task_queues[task_id]

        # 完成
        yield {
            "event": "done",
            "data": "{}"
        }

    return EventSourceResponse(event_generator())


@router.put("/api/video-records/{record_id}", response_model=VideoRecordResponse)
async def update_video_record(record_id: int, req: VideoRecordUpdate):
    """更新视频记录。"""
    with get_db() as conn:
        # 检查记录是否存在
        record = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        if not record:
            raise HTTPException(status_code=404, detail="记录不存在")

        # 构建更新语句
        updates = []
        params = []

        if req.video_time is not None:
            updates.append("video_time = ?")
            params.append(req.video_time)

        if req.category is not None:
            updates.append("category = ?")
            params.append(req.category)

        if req.product is not None:
            updates.append("product = ?")
            params.append(req.product)

        if req.golden_3s_copy is not None:
            updates.append("golden_3s_copy = ?")
            params.append(req.golden_3s_copy)

        if req.transcript is not None:
            updates.append("transcript = ?")
            params.append(req.transcript)

        if req.video_copy is not None:
            updates.append("video_copy = ?")
            params.append(req.video_copy)

        if req.viral_analysis is not None:
            updates.append("viral_analysis = ?")
            params.append(req.viral_analysis)

        if req.scene_analysis is not None:
            updates.append("scene_analysis = ?")
            params.append(req.scene_analysis)

        if req.exposure is not None:
            updates.append("exposure = ?")
            params.append(req.exposure)

        if req.likes is not None:
            updates.append("likes = ?")
            params.append(req.likes)

        if req.comments is not None:
            updates.append("comments = ?")
            params.append(req.comments)

        if req.shares is not None:
            updates.append("shares = ?")
            params.append(req.shares)

        if req.collects is not None:
            updates.append("collects = ?")
            params.append(req.collects)

        if req.remarks is not None:
            updates.append("remarks = ?")
            params.append(req.remarks)

        if updates:
            updates.append("updated_at = datetime('now')")
            params.append(record_id)

            sql = f"UPDATE video_records SET {', '.join(updates)} WHERE id = ?"
            conn.execute(sql, params)

        # 返回更新后的记录
        row = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        return dict(row)


@router.delete("/api/video-records/{record_id}")
async def delete_video_record(record_id: int):
    """删除视频记录。"""
    with get_db() as conn:
        # 检查记录是否存在
        record = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        if not record:
            raise HTTPException(status_code=404, detail="记录不存在")

        conn.execute("DELETE FROM video_records WHERE id = ?", (record_id,))

        return {"success": True, "message": "删除成功"}


@router.get("/api/video-records/{record_id}/play")
async def play_video(record_id: int):
    """播放视频，优先使用原视频，如果不存在则使用压缩视频。"""
    import os
    from fastapi.responses import FileResponse
    from app.services.video_compressor import compressor

    with get_db() as conn:
        record = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        if not record:
            raise HTTPException(status_code=404, detail="记录不存在")

        video_file_path = record["video_file_path"]

        # 检查原视频是否存在
        if video_file_path and os.path.exists(video_file_path):
            return FileResponse(
                path=video_file_path,
                media_type="video/mp4",
                filename=os.path.basename(video_file_path)
            )

        # 原视频不存在，尝试使用压缩视频
        if video_file_path:
            compressed_path = compressor.get_compressed_video(video_file_path)
            if compressed_path and os.path.exists(compressed_path):
                return FileResponse(
                    path=compressed_path,
                    media_type="video/mp4",
                    filename=os.path.basename(compressed_path)
                )

        raise HTTPException(status_code=404, detail="视频文件不存在")


@router.post("/api/video-records/{record_id}/download")
async def download_video_for_record(record_id: int):
    """手动下载视频到本地。"""
    with get_db() as conn:
        record = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        if not record:
            raise HTTPException(status_code=404, detail="记录不存在")

        video_url = record["video_url"]

        try:
            from app.services.video_service import download_video_service
            from app.schemas.video import DownloadRequest
            from app.integrations.douyin_client import is_douyin_url
            from app.integrations.yt_dlp_client import downloader

            # 重新解析获取最新的下载链接
            loop = asyncio.get_event_loop()
            # 抖音链接使用专用解析器（不需要 cookies）
            # 其他平台使用 yt-dlp
            if is_douyin_url(video_url):
                video_info = await loop.run_in_executor(None, douyin_parser.parse, video_url)
            else:
                video_info = await loop.run_in_executor(None, downloader.parse_video, video_url)

            formats = video_info.get("formats", [])
            format_id = "best"
            if formats and len(formats) > 0:
                format_id = formats[0].get("format_id", "best")

            download_req = DownloadRequest(url=video_url, format_id=format_id)
            download_result = await download_video_service(download_req)

            if download_result and "filepath" in download_result:
                video_file_path = download_result["filepath"]

                # 更新数据库中的视频路径
                conn.execute(
                    "UPDATE video_records SET video_file_path = ?, updated_at = datetime('now') WHERE id = ?",
                    (video_file_path, record_id)
                )

                return {"success": True, "message": "下载成功", "filepath": video_file_path}
            else:
                raise HTTPException(status_code=500, detail="下载失败")

        except Exception as e:
            raise HTTPException(status_code=400, detail={"success": False, "error": f"下载失败: {str(e)}"})


@router.post("/api/video-records/{record_id}/analyze")
async def analyze_video_with_gemini(record_id: int):
    """使用 Gemini 分析视频，提取口播文案和画面描述。"""
    from app.services.gemini_video_analyzer import analyzer

    with get_db() as conn:
        record = conn.execute(
            "SELECT * FROM video_records WHERE id = ?", (record_id,)
        ).fetchone()

        if not record:
            raise HTTPException(status_code=404, detail="记录不存在")

        if not analyzer.is_available():
            raise HTTPException(
                status_code=503,
                detail="Gemini 服务不可用，请在 api_config.yaml 中配置 Gemini API"
            )

        video_file_path = record["video_file_path"]
        if not video_file_path:
            raise HTTPException(status_code=400, detail="请先下载视频到本地")

        try:
            # 在后台执行分析
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None, analyzer.analyze_compressed_video, video_file_path
            )

            if not result:
                raise HTTPException(status_code=500, detail="视频分析失败")

            # 提取分析结果
            golden_3s = result.get("golden_3s", "")
            transcript = result.get("transcript", "")
            scenes = result.get("scenes", "")
            viral_analysis = result.get("viral_analysis", "")

            # 更新数据库：
            # - 黄金三秒保存到 golden_3s_copy
            # - 口播文案保存到 transcript（新字段，不覆盖原有 video_copy）
            # - 爆款分析保存到 viral_analysis（新字段）
            # - 画面描述保存到 remarks
            golden_3s_copy = golden_3s if golden_3s else record["golden_3s_copy"]
            transcript_text = transcript if transcript and transcript != "无口播" else record.get("transcript", "")
            viral_analysis_text = viral_analysis if viral_analysis else record.get("viral_analysis", "")
            remarks = f"【AI分析】\n画面: {scenes}"

            conn.execute(
                """UPDATE video_records
                   SET golden_3s_copy = ?, transcript = ?, viral_analysis = ?, remarks = ?, updated_at = datetime('now')
                   WHERE id = ?""",
                (golden_3s_copy, transcript_text, viral_analysis_text, remarks, record_id)
            )

            return {
                "success": True,
                "message": "分析完成",
                "data": {
                    "golden_3s": golden_3s,
                    "transcript": transcript,
                    "scenes": scenes,
                    "viral_analysis": viral_analysis,
                    "key_moments": result.get("key_moments", [])
                }
            }

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail={"success": False, "error": f"分析失败: {str(e)}"})


@router.post("/api/video-all-in-one")
async def video_all_in_one(video_url: str):
    """一体化接口：输入视频链接，自动完成解析、下载、压缩、AI分析，返回完整数据。

    Args:
        video_url: 视频链接

    Returns:
        完整的视频数据，包括：
        - 基本信息（标题、时长、平台等）
        - 互动数据（点赞、评论、分享、收藏）
        - 本地文件路径（原视频、压缩视频）
        - AI分析结果（口播文案、画面描述、爆款原因分析）
    """
    import os
    from app.services.gemini_video_analyzer import analyzer
    from app.integrations.douyin_client import is_douyin_url
    from app.integrations.yt_dlp_client import downloader
    from app.services.video_service import download_video_service
    from app.schemas.video import DownloadRequest
    from app.services.video_compressor import compressor

    try:
        # 1. 解析视频信息
        print(f"[一体化API] 步骤1: 解析视频信息")
        loop = asyncio.get_event_loop()

        if is_douyin_url(video_url):
            video_info = await loop.run_in_executor(None, douyin_parser.parse, video_url)
        else:
            video_info = await loop.run_in_executor(None, downloader.parse_video, video_url)

        # 提取基本信息
        video_data = {
            "video_url": video_url,
            "title": video_info.get("title", ""),
            "platform": video_info.get("platform", ""),
            "duration": video_info.get("duration", 0),
            "uploader": video_info.get("uploader", ""),
            "thumbnail": video_info.get("thumbnail", ""),
            "description": video_info.get("description", ""),
            "likes": video_info.get("like_count", 0) or 0,
            "comments": video_info.get("comment_count", 0) or 0,
            "shares": video_info.get("share_count", 0) or 0,
            "collects": video_info.get("collect_count", 0) or 0,
        }

        # 2. 下载视频
        print(f"[一体化API] 步骤2: 下载视频")
        formats = video_info.get("formats", [])
        format_id = "best"
        if formats and len(formats) > 0:
            format_id = formats[0].get("format_id", "best")

        download_req = DownloadRequest(url=video_url, format_id=format_id)
        download_result = await download_video_service(download_req)

        if not download_result or "filepath" not in download_result:
            raise HTTPException(status_code=500, detail="视频下载失败")

        video_file_path = download_result["filepath"]
        video_data["video_file_path"] = video_file_path
        video_data["video_filename"] = download_result.get("filename", "")

        # 3. 等待压缩完成（压缩是异步的，等待一下）
        print(f"[一体化API] 步骤3: 等待视频压缩")
        await asyncio.sleep(5)  # 等待压缩任务启动

        # 检查压缩视频
        compressed_path = compressor.get_compressed_video(video_file_path)
        if compressed_path:
            video_data["compressed_file_path"] = compressed_path
            video_data["compressed_filename"] = os.path.basename(compressed_path)
        else:
            video_data["compressed_file_path"] = None
            video_data["compressed_filename"] = None

        # 4. AI 分析（如果 Gemini 可用）
        if analyzer.is_available():
            print(f"[一体化API] 步骤4: AI 分析视频")
            try:
                # 等待压缩完成
                max_wait = 60  # 最多等待60秒
                waited = 0
                while waited < max_wait:
                    compressed_path = compressor.get_compressed_video(video_file_path)
                    if compressed_path and os.path.exists(compressed_path):
                        break
                    await asyncio.sleep(2)
                    waited += 2

                # 执行分析
                analysis_result = await loop.run_in_executor(
                    None, analyzer.analyze_compressed_video, video_file_path
                )

                if analysis_result:
                    video_data["ai_analysis"] = {
                        "golden_3s": analysis_result.get("golden_3s", ""),
                        "transcript": analysis_result.get("transcript", ""),
                        "scenes": analysis_result.get("scenes", ""),
                        "viral_analysis": analysis_result.get("viral_analysis", ""),
                        "key_moments": analysis_result.get("key_moments", [])
                    }
                else:
                    video_data["ai_analysis"] = None
            except Exception as e:
                print(f"[一体化API] AI分析失败: {e}")
                video_data["ai_analysis"] = {"error": str(e)}
        else:
            video_data["ai_analysis"] = None

        print(f"[一体化API] 完成！")
        return {
            "success": True,
            "message": "视频处理完成",
            "data": video_data
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail={"success": False, "error": str(e)})


@router.get("/api/video-records/{record_id}/stream")
async def stream_ai_updates(record_id: int):
    """SSE 端点：实时推送 AI 分析结果更新。"""
    import queue

    print(f"[SSE连接] ========== 记录 {record_id} 建立SSE连接 ==========", flush=True)

    # 为这个记录创建一个队列
    update_queue = queue.Queue()
    ai_update_queues[record_id] = update_queue
    print(f"[SSE连接] ✓ 已为记录 {record_id} 创建更新队列", flush=True)
    print(f"[SSE连接] 当前所有队列: {list(ai_update_queues.keys())}", flush=True)

    async def event_generator():
        try:
            # 先检查数据库中是否已有 AI 分析结果
            with get_db() as conn:
                record = conn.execute(
                    "SELECT golden_3s_copy, transcript, viral_analysis, remarks FROM video_records WHERE id = ?",
                    (record_id,)
                ).fetchone()

                if record:
                    # 如果已有数据，立即推送
                    if record["golden_3s_copy"] or record["transcript"] or record["viral_analysis"]:
                        print(f"[SSE] 记录 {record_id} 已有 AI 数据，立即推送")
                        yield {
                            "event": "ai_update",
                            "data": json.dumps({
                                "golden_3s_copy": record["golden_3s_copy"] or "",
                                "transcript": record["transcript"] or "",
                                "viral_analysis": record["viral_analysis"] or "",
                                "remarks": record["remarks"] or ""
                            }, ensure_ascii=False)
                        }
                        # 已有数据，关闭连接
                        yield {"event": "done", "data": "{}"}
                        return

            print(f"[SSE] 记录 {record_id} 等待 AI 分析完成...")
            # 等待 AI 分析完成（最多等待 10 分钟）
            timeout = 600
            start_time = asyncio.get_event_loop().time()

            while True:
                # 检查超时
                if asyncio.get_event_loop().time() - start_time > timeout:
                    print(f"[SSE] 记录 {record_id} 等待超时")
                    yield {"event": "timeout", "data": "{}"}
                    break

                # 非阻塞地检查队列
                try:
                    update_data = update_queue.get_nowait()
                    print(f"[SSE] 记录 {record_id} 收到 AI 更新，准备推送: {update_data}")
                    # 推送更新
                    yield {
                        "event": "ai_update",
                        "data": json.dumps(update_data, ensure_ascii=False)
                    }
                    # 推送完成，关闭连接
                    yield {"event": "done", "data": "{}"}
                    print(f"[SSE] 记录 {record_id} 推送完成，关闭连接")
                    break
                except queue.Empty:
                    # 队列为空，等待一会儿再检查
                    await asyncio.sleep(1)

        finally:
            # 清理队列
            if record_id in ai_update_queues:
                del ai_update_queues[record_id]
                print(f"[SSE] 已清理记录 {record_id} 的更新队列")

    return EventSourceResponse(event_generator())


@router.get("/api/video-records/export")
async def export_to_excel(sheet_id: str = "sheet1"):
    """导出当前工作表的视频记录为Excel文件"""
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse

        # 格式化数字为 K/W 形式
        def format_number(num):
            if num is None or num == '':
                return ''
            try:
                num = int(num)
                if num >= 10000:
                    return f"{num / 10000:.1f}W"
                elif num >= 1000:
                    return f"{num / 1000:.1f}K"
                else:
                    return str(num)
            except (ValueError, TypeError):
                return str(num)

        # 如果 sheet_id 是数字字符串，转换为整数
        try:
            sheet_id_value = int(sheet_id)
        except (ValueError, TypeError):
            sheet_id_value = sheet_id

        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT video_url, video_time, category, product,
                       golden_3s_copy, transcript, copywriting, video_copy, viral_analysis,
                       exposure, likes, comments, shares, collects, remarks
                FROM video_records
                WHERE sheet_id = ?
                ORDER BY id DESC
            """, (sheet_id_value,))
            records = cursor.fetchall()

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "视频数据"

        # 设置表头（去掉序号列）
        headers = ["视频链接", "视频时间", "品类", "产品",
                   "黄金三秒文案", "口播文案", "文案仿写", "视频文案", "爆款分析",
                   "曝光量", "点赞量", "评论数", "分享数", "收藏数", "备注"]

        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 边框样式
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 数据单元格样式
        data_alignment = Alignment(vertical="top", wrap_text=True)
        data_font = Font(size=10)

        # 写入表头并设置样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置表头行高
        ws.row_dimensions[1].height = 30

        # 写入数据
        for row_num, record in enumerate(records, 2):
            for col_num, value in enumerate(record, 1):
                # 格式化数字列（曝光量、点赞量、评论数、分享数、收藏数）
                if col_num >= 10 and col_num <= 14:
                    value = format_number(value)

                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = data_alignment
                cell.font = data_font
                cell.border = thin_border

                # 数字列右对齐
                if col_num >= 10 and col_num <= 14:  # 曝光量到收藏数
                    cell.alignment = Alignment(horizontal="right", vertical="top")

            # 设置数据行高
            ws.row_dimensions[row_num].height = 60

        # 调整列宽（去掉序号列后的宽度）
        column_widths = [45, 12, 12, 18, 35, 45, 45, 45, 45, 12, 10, 10, 10, 10, 30]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=video_data_{sheet_id}.xlsx"}
        )
    except Exception as e:
        print(f"[导出Excel错误] {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/api/video-records/{record_id}/upload-video")
async def upload_video_manually(record_id: int, file: UploadFile = File(...)):
    """
    手动上传视频并触发压缩和AI分析

    用于解析失败的视频，用户手动上传后重新分析
    """
    from pathlib import Path
    import shutil
    import os

    try:
        print(f"[手动上传] 记录 {record_id} 开始上传视频: {file.filename}", flush=True)

        # 1. 验证记录是否存在
        with get_db() as conn:
            row = conn.execute(
                "SELECT * FROM video_records WHERE id = ?", (record_id,)
            ).fetchone()

            if not row:
                raise HTTPException(status_code=404, detail="视频记录不存在")

            # 检查是否已有视频文件（防止重复上传）
            if row['video_file_path'] and os.path.exists(row['video_file_path']):
                # 检查文件是否是手动上传的（允许重新上传）
                if 'manual_upload' not in row['video_file_path']:
                    raise HTTPException(status_code=400, detail="该记录已有视频文件，无需重复上传")

        # 2. 验证文件类型（只允许视频格式）
        allowed_extensions = {'.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.webm'}
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件格式: {file_ext}，仅支持视频文件 ({', '.join(allowed_extensions)})"
            )

        # 3. 验证文件大小（限制500MB）
        MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB
        file.file.seek(0, 2)  # 移动到文件末尾
        file_size = file.file.tell()
        file.file.seek(0)  # 重置到开头

        if file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"文件过大: {file_size / 1024 / 1024:.1f}MB，最大允许 500MB"
            )

        if file_size < 1024:  # 小于1KB
            raise HTTPException(status_code=400, detail="文件过小，可能不是有效的视频文件")

        # 4. 检查磁盘空间（至少需要文件大小的2倍空间用于压缩）
        download_dir = Path(__file__).parent.parent.parent.parent / "downloads"
        download_dir.mkdir(parents=True, exist_ok=True)

        stat = os.statvfs(str(download_dir)) if hasattr(os, 'statvfs') else None
        if stat:
            free_space = stat.f_bavail * stat.f_frsize
            required_space = file_size * 2
            if free_space < required_space:
                raise HTTPException(
                    status_code=507,
                    detail=f"磁盘空间不足，需要 {required_space / 1024 / 1024:.1f}MB，可用 {free_space / 1024 / 1024:.1f}MB"
                )

        # 5. 保存上传的视频文件
        safe_filename = f"manual_upload_{record_id}_{int(time.time())}{file_ext}"
        filepath = download_dir / safe_filename

        print(f"[手动上传] 保存文件: {filepath} ({file_size / 1024 / 1024:.2f}MB)", flush=True)

        # 分块写入，避免大文件内存溢出
        CHUNK_SIZE = 1024 * 1024  # 1MB chunks
        with open(filepath, "wb") as buffer:
            while True:
                chunk = file.file.read(CHUNK_SIZE)
                if not chunk:
                    break
                buffer.write(chunk)

        print(f"[手动上传] 视频已保存: {filepath}", flush=True)

        # 6. 验证文件完整性（检查文件大小）
        actual_size = os.path.getsize(filepath)
        if actual_size != file_size:
            os.remove(filepath)  # 删除不完整的文件
            raise HTTPException(
                status_code=500,
                detail=f"文件上传不完整，预期 {file_size} 字节，实际 {actual_size} 字节"
            )

        # 7. 更新数据库
        with get_db() as conn:
            conn.execute(
                "UPDATE video_records SET video_file_path = ?, updated_at = datetime('now') WHERE id = ?",
                (str(filepath), record_id)
            )

        # 8. 启动压缩和AI分析
        loop = asyncio.get_event_loop()

        def compress_and_analyze(vid_path: str, rec_id: int):
            try:
                from app.services.video_compressor import compressor
                from app.services.gemini_video_analyzer import analyzer

                print(f"[手动上传 {rec_id}] 开始压缩视频", flush=True)
                compress_result = compressor.compress_video(vid_path, "medium", 1280)

                if compress_result:
                    print(f"[手动上传 {rec_id}] 压缩完成，开始 AI 分析", flush=True)
                else:
                    print(f"[手动上传 {rec_id}] 压缩失败，使用原视频进行 AI 分析", flush=True)

                if analyzer.is_available():
                    print(f"[手动上传 {rec_id}] 开始 AI 分析视频", flush=True)
                    analysis_result = analyzer.analyze_compressed_video(vid_path)

                    if analysis_result:
                        print(f"[手动上传 {rec_id}] AI 分析完成，更新数据库", flush=True)
                        with get_db() as conn:
                            category = analysis_result.get("category", "")
                            product = analysis_result.get("product", "")
                            golden_3s = analysis_result.get("golden_3s", "")
                            transcript = analysis_result.get("transcript", "")
                            viral_analysis = analysis_result.get("viral_analysis", "")
                            scenes = analysis_result.get("scenes", "")

                            conn.execute(
                                """UPDATE video_records
                                   SET category = ?, product = ?, golden_3s_copy = ?,
                                       transcript = ?, viral_analysis = ?, scene_analysis = ?,
                                       updated_at = datetime('now')
                                   WHERE id = ?""",
                                (category, product, golden_3s, transcript, viral_analysis, scenes, rec_id)
                            )
                        print(f"[手动上传 {rec_id}] AI 分析结果已保存", flush=True)
                    else:
                        print(f"[手动上传 {rec_id}] AI 分析失败", flush=True)
                else:
                    print(f"[手动上传 {rec_id}] Gemini 服务不可用，跳过 AI 分析", flush=True)

            except Exception as e:
                print(f"[手动上传 {rec_id}] 压缩和分析任务失败: {e}", flush=True)
                import traceback
                traceback.print_exc()

        # 在后台执行
        loop.run_in_executor(None, compress_and_analyze, str(filepath), record_id)

        return {
            "success": True,
            "message": "视频上传成功，正在后台进行AI分析",
            "record_id": record_id,
            "video_path": str(filepath),
            "file_size_mb": round(file_size / 1024 / 1024, 2)
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"[手动上传] 上传失败: {e}", flush=True)
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")



